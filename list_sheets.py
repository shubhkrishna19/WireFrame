import openpyxl

# Load the Excel file
file_path = 'fullreport.xlsm'
workbook = openpyxl.load_workbook(file_path, data_only=True)

# List all sheet names
print("Sheet Names:")
for idx, sheet_name in enumerate(workbook.sheetnames, 1):
    sheet = workbook[sheet_name]
    print(f"{idx}. {sheet_name} - Rows: {sheet.max_row}, Cols: {sheet.max_column}")
