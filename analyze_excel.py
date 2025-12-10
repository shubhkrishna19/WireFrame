import openpyxl
import pandas as pd

# Load the Excel file
file_path = 'fullreport.xlsm'
workbook = openpyxl.load_workbook(file_path, data_only=True)

# List all sheet names
print("=" * 80)
print("ALL SHEETS IN fullreport.xlsm:")
print("=" * 80)
for idx, sheet_name in enumerate(workbook.sheetnames, 1):
    print(f"{idx}. {sheet_name}")

print("\n" + "=" * 80)
print("ANALYZING EACH SHEET:")
print("=" * 80)

# Analyze each sheet
for sheet_name in workbook.sheetnames:
    print(f"\n\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    
    try:
        # Read the sheet with pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        
        print(f"Rows: {len(df)}")
        print(f"Columns: {len(df.columns)}")
        print(f"\nColumn Names:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")
        
        # Show first 2 rows
        print(f"\nFirst 2 rows preview:")
        print(df.head(2).to_string())
        
    except Exception as e:
        print(f"Error reading sheet: {e}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
